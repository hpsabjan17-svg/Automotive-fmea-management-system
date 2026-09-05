from flask import Flask, render_template, request, redirect, url_for, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os

app = Flask(__name__)

# =========================================================
# SECURITY / SESSION
# =========================================================

app.secret_key = os.environ.get(
    "SECRET_KEY",
    "development-secret-key-change-in-render"
)

app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = (
    os.environ.get("COOKIE_SECURE", "false").lower() == "true"
)

DATABASE = os.path.join("database", "fmea.db")


# =========================================================
# DATABASE
# =========================================================

def get_db():
    os.makedirs("database", exist_ok=True)

    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row

    return conn


def add_column_if_missing(
    cursor,
    table_name,
    column_name,
    column_definition
):
    columns = cursor.execute(
        f"PRAGMA table_info({table_name})"
    ).fetchall()

    existing = [column[1] for column in columns]

    if column_name not in existing:
        cursor.execute(
            f"""
            ALTER TABLE {table_name}
            ADD COLUMN {column_name} {column_definition}
            """
        )


def setup_database():

    conn = get_db()
    cursor = conn.cursor()

    # =====================================================
    # USERS
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_date TEXT NOT NULL
        )
    """)

    admin_username = os.environ.get("ADMIN_USERNAME")
    admin_password = os.environ.get("ADMIN_PASSWORD")

    if admin_username and admin_password:

        existing = cursor.execute(
            """
            SELECT id
            FROM users
            WHERE username = ?
            """,
            (admin_username,)
        ).fetchone()

        if existing is None:

            cursor.execute("""
                INSERT INTO users
                (
                    username,
                    password_hash,
                    created_date
                )
                VALUES (?, ?, ?)
            """, (
                admin_username,
                generate_password_hash(admin_password),
                date.today().isoformat()
            ))

    # =====================================================
    # PROJECTS
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT,
            product_name TEXT,
            customer TEXT,
            project_number TEXT,
            created_date TEXT
        )
    """)

    for column, definition in [
        ("project_name", "TEXT"),
        ("product_name", "TEXT"),
        ("customer", "TEXT"),
        ("project_number", "TEXT"),
        ("created_date", "TEXT"),
        ("oem_name", "TEXT DEFAULT 'Generic'"),
        ("compliance_mode", "TEXT DEFAULT 'AIAG-VDA 2019'")
    ]:
        add_column_if_missing(
            cursor,
            "projects",
            column,
            definition
        )

    # =====================================================
    # OEM STANDARDS
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS oem_standards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            oem_name TEXT NOT NULL UNIQUE,
            standard_framework TEXT NOT NULL,
            cc_symbol TEXT,
            sc_symbol TEXT,
            archiving_period_years INTEGER DEFAULT 10,
            description TEXT
        )
    """)

    oem_data = [
        (
            "Volkswagen Group",
            "AIAG-VDA",
            "D/TLD",
            "K",
            15,
            "OEM CSR prototype configuration for Volkswagen Group."
        ),
        (
            "BMW Group",
            "AIAG-VDA",
            "DS",
            "PTC",
            12,
            "OEM CSR prototype configuration for BMW Group."
        ),
        (
            "Ford Motor Co",
            "AIAG-VDA",
            "∇",
            "SC",
            10,
            "OEM CSR prototype configuration for Ford Motor Co."
        ),
        (
            "General Motors",
            "AIAG-VDA",
            "KPC",
            "PQC",
            10,
            "OEM CSR prototype configuration for General Motors."
        ),
        (
            "Stellantis",
            "AIAG-VDA",
            "S",
            "R",
            10,
            "OEM CSR prototype configuration for Stellantis."
        ),
        (
            "Generic",
            "AIAG-VDA 2019",
            "CC",
            "SC",
            10,
            "Generic FMEA configuration."
        )
    ]

    cursor.executemany("""
        INSERT OR IGNORE INTO oem_standards
        (
            oem_name,
            standard_framework,
            cc_symbol,
            sc_symbol,
            archiving_period_years,
            description
        )
        VALUES (?, ?, ?, ?, ?, ?)
    """, oem_data)

    # =====================================================
    # FUNCTIONAL ANALYSIS
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS functional_analysis (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            level TEXT,
            surrounding_assembly TEXT,
            function TEXT,
            requirement TEXT
        )
    """)

    for column, definition in [
        ("project_id", "INTEGER"),
        ("level", "TEXT"),
        ("surrounding_assembly", "TEXT"),
        ("function", "TEXT"),
        ("requirement", "TEXT")
    ]:
        add_column_if_missing(
            cursor,
            "functional_analysis",
            column,
            definition
        )

    # =====================================================
    # BOUNDARY DIAGRAM
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS boundary_diagram (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            external_element TEXT,
            interaction TEXT,
            direction TEXT,
            description TEXT
        )
    """)

    for column, definition in [
        ("project_id", "INTEGER"),
        ("external_element", "TEXT"),
        ("interaction", "TEXT"),
        ("direction", "TEXT"),
        ("description", "TEXT")
    ]:
        add_column_if_missing(
            cursor,
            "boundary_diagram",
            column,
            definition
        )

    # =====================================================
    # PRODUCT STRUCTURE
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS product_structure (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            parent_id INTEGER,
            component_name TEXT,
            component_type TEXT,
            label TEXT,
            part_number TEXT,
            level INTEGER DEFAULT 0,
            description TEXT
        )
    """)

    for column, definition in [
        ("project_id", "INTEGER"),
        ("parent_id", "INTEGER"),
        ("component_name", "TEXT"),
        ("component_type", "TEXT"),
        ("label", "TEXT"),
        ("part_number", "TEXT"),
        ("level", "INTEGER DEFAULT 0"),
        ("description", "TEXT")
    ]:
        add_column_if_missing(
            cursor,
            "product_structure",
            column,
            definition
        )

    # =====================================================
    # KEY CHARACTERISTICS
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS key_characteristics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            component_id INTEGER,
            characteristic TEXT,
            specification TEXT,
            tolerance TEXT,
            severity INTEGER,
            responsibility TEXT
        )
    """)

    for column, definition in [
        ("project_id", "INTEGER"),
        ("component_id", "INTEGER"),
        ("characteristic", "TEXT"),
        ("specification", "TEXT"),
        ("tolerance", "TEXT"),
        ("severity", "INTEGER"),
        ("responsibility", "TEXT")
    ]:
        add_column_if_missing(
            cursor,
            "key_characteristics",
            column,
            definition
        )

    # =====================================================
    # FUNCTIONAL LINKS
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS functional_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            function_id INTEGER,
            component_id INTEGER,
            requirement TEXT
        )
    """)

    for column, definition in [
        ("project_id", "INTEGER"),
        ("function_id", "INTEGER"),
        ("component_id", "INTEGER"),
        ("requirement", "TEXT")
    ]:
        add_column_if_missing(
            cursor,
            "functional_links",
            column,
            definition
        )

    # =====================================================
    # DFMEA
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS dfmea (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            component_id INTEGER,
            function TEXT,
            failure_mode TEXT,
            failure_effect TEXT,
            severity INTEGER,
            cause TEXT,
            occurrence INTEGER,
            prevention_control TEXT,
            detection_control TEXT,
            detection INTEGER,
            rpn INTEGER,
            recommended_action TEXT,
            responsibility TEXT,
            target_date TEXT,
            action_status TEXT
        )
    """)

    for column, definition in [
        ("project_id", "INTEGER"),
        ("component_id", "INTEGER"),
        ("function", "TEXT"),
        ("failure_mode", "TEXT"),
        ("failure_effect", "TEXT"),
        ("severity", "INTEGER"),
        ("cause", "TEXT"),
        ("occurrence", "INTEGER"),
        ("prevention_control", "TEXT"),
        ("detection_control", "TEXT"),
        ("detection", "INTEGER"),
        ("rpn", "INTEGER"),
        ("recommended_action", "TEXT"),
        ("responsibility", "TEXT"),
        ("target_date", "TEXT"),
        ("action_status", "TEXT")
    ]:
        add_column_if_missing(
            cursor,
            "dfmea",
            column,
            definition
        )

    # =====================================================
    # PFMEA
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS pfmea (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            component_id INTEGER,
            process_step TEXT,
            process_function TEXT,
            failure_mode TEXT,
            failure_effect TEXT,
            severity INTEGER,
            cause TEXT,
            occurrence INTEGER,
            prevention_control TEXT,
            detection_control TEXT,
            detection INTEGER,
            rpn INTEGER,
            recommended_action TEXT,
            responsibility TEXT,
            target_date TEXT,
            action_status TEXT
        )
    """)

    for column, definition in [
        ("project_id", "INTEGER"),
        ("component_id", "INTEGER"),
        ("process_step", "TEXT"),
        ("process_function", "TEXT"),
        ("failure_mode", "TEXT"),
        ("failure_effect", "TEXT"),
        ("severity", "INTEGER"),
        ("cause", "TEXT"),
        ("occurrence", "INTEGER"),
        ("prevention_control", "TEXT"),
        ("detection_control", "TEXT"),
        ("detection", "INTEGER"),
        ("rpn", "INTEGER"),
        ("recommended_action", "TEXT"),
        ("responsibility", "TEXT"),
        ("target_date", "TEXT"),
        ("action_status", "TEXT")
    ]:
        add_column_if_missing(
            cursor,
            "pfmea",
            column,
            definition
        )

    # =====================================================
    # CONTROL PLAN
    # =====================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS control_plan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            component_id INTEGER,
            process_step TEXT,
            characteristic TEXT,
            specification TEXT,
            control_method TEXT,
            measurement_method TEXT,
            sample_size TEXT,
            frequency TEXT,
            responsibility TEXT,
            reaction_plan TEXT
        )
    """)

    for column, definition in [
        ("project_id", "INTEGER"),
        ("component_id", "INTEGER"),
        ("process_step", "TEXT"),
        ("characteristic", "TEXT"),
        ("specification", "TEXT"),
        ("control_method", "TEXT"),
        ("measurement_method", "TEXT"),
        ("sample_size", "TEXT"),
        ("frequency", "TEXT"),
        ("responsibility", "TEXT"),
        ("reaction_plan", "TEXT")
    ]:
        add_column_if_missing(
            cursor,
            "control_plan",
            column,
            definition
        )

    conn.commit()
    conn.close()


# =========================================================
# LOGIN PROTECTION
# =========================================================

@app.before_request
def require_login():

    if request.endpoint in {"login", "static"}:
        return

    if "user_id" not in session:
        return redirect(url_for("login"))


# =========================================================
# LOGIN
# =========================================================

@app.route("/login", methods=["GET", "POST"])
def login():

    if session.get("user_id"):
        return redirect(url_for("dashboard"))

    error = None

    if request.method == "POST":

        username = request.form.get(
            "username",
            ""
        ).strip()

        password = request.form.get(
            "password",
            ""
        )

        conn = get_db()

        user = conn.execute(
            """
            SELECT *
            FROM users
            WHERE username = ?
            """,
            (username,)
        ).fetchone()

        conn.close()

        if user and check_password_hash(
            user["password_hash"],
            password
        ):

            session.clear()

            session["user_id"] = user["id"]
            session["username"] = user["username"]

            return redirect(
                url_for("dashboard")
            )

        error = "Invalid username or password."

    return render_template(
        "login.html",
        error=error
    )


# =========================================================
# LOGOUT
# =========================================================

@app.route("/logout")
def logout():

    session.clear()

    return redirect(
        url_for("login")
    )


# =========================================================
# PAGE FLOW
# =========================================================

PAGE_FLOW = [
    ("project", "Project"),
    ("functional_analysis", "Functional Analysis"),
    ("boundary_diagram", "Boundary Diagram"),
    ("product_structure", "Product Structure"),
    ("key_characteristics", "Key Characteristics"),
    ("functional_links", "Functional Links"),
    ("dfmea", "DFMEA"),
    ("pfmea", "PFMEA"),
    ("control_plan", "Control Plan"),
    ("reports", "Reports")
]


def get_next_page(current_endpoint):

    for index, (endpoint, title) in enumerate(PAGE_FLOW):

        if endpoint == current_endpoint:

            if index < len(PAGE_FLOW) - 1:

                return PAGE_FLOW[index + 1]

            return None

    return None


def get_previous_page(current_endpoint):

    for index, (endpoint, title) in enumerate(PAGE_FLOW):

        if endpoint == current_endpoint:

            if index > 0:

                return PAGE_FLOW[index - 1]

            return None

    return None


def page_navigation(current_endpoint, project_id=None):

    next_page = get_next_page(
        current_endpoint
    )

    previous_page = get_previous_page(
        current_endpoint
    )

    next_url = None
    previous_url = None

    if next_page:

        next_endpoint = next_page[0]

        if project_id:
            next_url = url_for(
                next_endpoint,
                project_id=project_id
            )
        else:
            next_url = url_for(
                next_endpoint
            )

    if previous_page:

        previous_endpoint = previous_page[0]

        if project_id:
            previous_url = url_for(
                previous_endpoint,
                project_id=project_id
            )
        else:
            previous_url = url_for(
                previous_endpoint
            )

    return {
        "next_url": next_url,
        "next_title": next_page[1] if next_page else None,
        "previous_url": previous_url,
        "previous_title": (
            previous_page[1]
            if previous_page
            else None
        )
    }


# =========================================================
# DASHBOARD
# =========================================================

@app.route("/")
def dashboard():

    conn = get_db()

    projects = conn.execute(
        """
        SELECT *
        FROM projects
        ORDER BY id DESC
        """
    ).fetchall()

    conn.close()

    return render_template(
        "dashboard.html",
        projects=projects
    )


# =========================================================
# PROJECT
# =========================================================

@app.route("/project", methods=["GET", "POST"])
def project():

    if request.method == "POST":

        project_name = request.form.get(
            "project_name",
            ""
        ).strip()

        product_name = request.form.get(
            "product_name",
            ""
        ).strip()

        customer = request.form.get(
            "customer",
            ""
        ).strip()

        oem_name = request.form.get(
            "oem_name",
            "Generic"
        ).strip()

        compliance_mode = request.form.get(
            "compliance_mode",
            "AIAG-VDA 2019"
        ).strip()

        project_number = request.form.get(
            "project_number",
            ""
        ).strip()

        created_date = request.form.get(
            "created_date",
            ""
        ).strip()

        if not created_date:
            created_date = date.today().isoformat()

        if project_name:

            conn = get_db()

            cursor = conn.cursor()

            cursor.execute(
                """
                INSERT INTO projects
                (
                    project_name,
                    product_name,
                    customer,
                    oem_name,
                    compliance_mode,
                    project_number,
                    created_date
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_name,
                    product_name,
                    customer,
                    oem_name,
                    compliance_mode,
                    project_number,
                    created_date
                )
            )

            project_id = cursor.lastrowid

            conn.commit()
            conn.close()

            return redirect(
                url_for(
                    "functional_analysis",
                    project_id=project_id
                )
            )

    conn = get_db()

    projects = conn.execute(
        """
        SELECT *
        FROM projects
        ORDER BY id DESC
        """
    ).fetchall()

    conn.close()

    navigation = page_navigation(
        "project"
    )

    return render_template(
        "project.html",
        projects=projects,
        today=date.today().isoformat(),
        **navigation
    )


# =========================================================
# FUNCTIONAL ANALYSIS
# =========================================================

@app.route(
    "/functional-analysis",
    methods=["GET", "POST"]
)
def functional_analysis():

    conn = get_db()

    selected_project_id = request.args.get(
        "project_id",
        ""
    )

    if request.method == "POST":

        project_id = request.form.get(
            "project_id",
            ""
        )

        level = request.form.get(
            "level",
            ""
        ).strip()

        surrounding_assembly = request.form.get(
            "surrounding_assembly",
            ""
        ).strip()

        function = request.form.get(
            "function",
            ""
        ).strip()

        requirement = request.form.get(
            "requirement",
            ""
        ).strip()

        if project_id and function:

            conn.execute(
                """
                INSERT INTO functional_analysis
                (
                    project_id,
                    level,
                    surrounding_assembly,
                    function,
                    requirement
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    level,
                    surrounding_assembly,
                    function,
                    requirement
                )
            )

            conn.commit()

            selected_project_id = project_id

    projects = conn.execute(
        """
        SELECT *
        FROM projects
        ORDER BY project_name
        """
    ).fetchall()

    records = conn.execute(
        """
        SELECT
            fa.id,
            fa.project_id,
            p.project_name,
            fa.level,
            fa.surrounding_assembly,
            fa.function,
            fa.requirement
        FROM functional_analysis AS fa
        LEFT JOIN projects AS p
            ON fa.project_id = p.id
        ORDER BY fa.id DESC
        """
    ).fetchall()

    conn.close()

    navigation = page_navigation(
        "functional_analysis",
        selected_project_id
    )

    return render_template(
        "functional_analysis.html",
        projects=projects,
        records=records,
        selected_project_id=selected_project_id,
        **navigation
    )


# =========================================================
# BOUNDARY DIAGRAM
# =========================================================

@app.route(
    "/boundary-diagram",
    methods=["GET", "POST"]
)
def boundary_diagram():

    conn = get_db()

    selected_project_id = request.args.get(
        "project_id",
        ""
    )

    if request.method == "POST":

        project_id = request.form.get(
            "project_id",
            ""
        )

        external_element = request.form.get(
            "external_element",
            ""
        ).strip()

        interaction = request.form.get(
            "interaction",
            ""
        ).strip()

        direction = request.form.get(
            "direction",
            ""
        ).strip()

        description = request.form.get(
            "description",
            ""
        ).strip()

        if project_id and external_element:

            conn.execute(
                """
                INSERT INTO boundary_diagram
                (
                    project_id,
                    external_element,
                    interaction,
                    direction,
                    description
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    external_element,
                    interaction,
                    direction,
                    description
                )
            )

            conn.commit()

            selected_project_id = project_id

    projects = conn.execute(
        """
        SELECT *
        FROM projects
        ORDER BY project_name
        """
    ).fetchall()

    boundaries = conn.execute(
        """
        SELECT
            bd.id,
            bd.project_id,
            bd.external_element,
            bd.interaction,
            bd.direction,
            bd.description,
            p.project_name
        FROM boundary_diagram AS bd
        LEFT JOIN projects AS p
            ON bd.project_id = p.id
        ORDER BY bd.id DESC
        """
    ).fetchall()

    conn.close()

    navigation = page_navigation(
        "boundary_diagram",
        selected_project_id
    )

    return render_template(
        "boundary_diagram.html",
        projects=projects,
        boundaries=boundaries,
        selected_project_id=selected_project_id,
        **navigation
    )


# =========================================================
# PRODUCT STRUCTURE
# =========================================================

@app.route(
    "/product-structure",
    methods=["GET", "POST"]
)
def product_structure():

    conn = get_db()

    selected_project_id = request.args.get(
        "project_id",
        ""
    )

    if request.method == "POST":

        project_id = request.form.get(
            "project_id",
            ""
        )

        parent_id = request.form.get(
            "parent_id",
            ""
        )

        component_name = request.form.get(
            "component_name",
            ""
        ).strip()

        component_type = request.form.get(
            "component_type",
            ""
        ).strip()

        label = request.form.get(
            "label",
            ""
        ).strip()

        part_number = request.form.get(
            "part_number",
            ""
        ).strip()

        level = request.form.get(
            "level",
            "0"
        ).strip()

        description = request.form.get(
            "description",
            ""
        ).strip()

        if parent_id == "":
            parent_id = None

        if project_id and component_name:

            conn.execute(
                """
                INSERT INTO product_structure
                (
                    project_id,
                    parent_id,
                    component_name,
                    component_type,
                    label,
                    part_number,
                    level,
                    description
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    parent_id,
                    component_name,
                    component_type,
                    label,
                    part_number,
                    level,
                    description
                )
            )

            conn.commit()

            conn.close()

            return redirect(
                url_for(
                    "product_structure",
                    project_id=project_id
                )
            )

    projects = conn.execute(
        """
        SELECT
            id,
            project_name,
            product_name
        FROM projects
        ORDER BY id DESC
        """
    ).fetchall()

    if selected_project_id:

        components = conn.execute(
            """
            SELECT *
            FROM product_structure
            WHERE project_id = ?
            ORDER BY level, id
            """,
            (selected_project_id,)
        ).fetchall()

    else:
        components = []

    conn.close()

    navigation = page_navigation(
        "product_structure",
        selected_project_id
    )

    return render_template(
        "product_structure.html",
        projects=projects,
        components=components,
        selected_project_id=selected_project_id,
        **navigation
    )


# =========================================================
# KEY CHARACTERISTICS
# =========================================================

@app.route(
    "/key-characteristics",
    methods=["GET", "POST"]
)
def key_characteristics():

    conn = get_db()

    selected_project_id = request.args.get(
        "project_id",
        ""
    )

    if request.method == "POST":

        project_id = request.form.get(
            "project_id",
            ""
        )

        component_id = request.form.get(
            "component_id",
            ""
        )

        characteristic = request.form.get(
            "characteristic",
            ""
        ).strip()

        specification = request.form.get(
            "specification",
            ""
        ).strip()

        tolerance = request.form.get(
            "tolerance",
            ""
        ).strip()

        severity = request.form.get(
            "severity",
            "1"
        )

        responsibility = request.form.get(
            "responsibility",
            ""
        ).strip()

        if project_id and component_id and characteristic:

            conn.execute(
                """
                INSERT INTO key_characteristics
                (
                    project_id,
                    component_id,
                    characteristic,
                    specification,
                    tolerance,
                    severity,
                    responsibility
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    component_id,
                    characteristic,
                    specification,
                    tolerance,
                    severity,
                    responsibility
                )
            )

            conn.commit()

            conn.close()

            return redirect(
                url_for(
                    "key_characteristics",
                    project_id=project_id
                )
            )

    projects = conn.execute(
        """
        SELECT
            id,
            project_name,
            product_name
        FROM projects
        ORDER BY id DESC
        """
    ).fetchall()

    if selected_project_id:

        components = conn.execute(
            """
            SELECT
                id,
                project_id,
                component_name,
                component_type,
                label,
                part_number,
                level
            FROM product_structure
            WHERE project_id = ?
            ORDER BY level, id
            """,
            (selected_project_id,)
        ).fetchall()

    else:
        components = []

    records = conn.execute(
        """
        SELECT
            kc.id,
            kc.project_id,
            kc.component_id,
            p.project_name,
            ps.component_name,
            kc.characteristic,
            kc.specification,
            kc.tolerance,
            kc.severity,
            kc.responsibility
        FROM key_characteristics AS kc
        LEFT JOIN projects AS p
            ON kc.project_id = p.id
        LEFT JOIN product_structure AS ps
            ON kc.component_id = ps.id
        ORDER BY kc.id DESC
        """
    ).fetchall()

    conn.close()

    navigation = page_navigation(
        "key_characteristics",
        selected_project_id
    )

    return render_template(
        "key_characteristics.html",
        projects=projects,
        components=components,
        records=records,
        selected_project_id=selected_project_id,
        **navigation
    )

# =========================================================
# DFMEA
# =========================================================

@app.route("/dfmea", methods=["GET", "POST"])
def dfmea():
    conn = get_db()
    selected_project_id = request.args.get("project_id", "")

    if request.method == "POST":
        project_id = request.form.get("project_id", "")
        component_id = request.form.get("component_id", "")

        function = request.form.get("function", "").strip()
        failure_mode = request.form.get("failure_mode", "").strip()
        failure_effect = request.form.get("failure_effect", "").strip()

        severity = request.form.get("severity", "1")
        cause = request.form.get("cause", "").strip()
        occurrence = request.form.get("occurrence", "1")

        prevention_control = request.form.get(
            "prevention_control", ""
        ).strip()

        detection_control = request.form.get(
            "detection_control", ""
        ).strip()

        detection = request.form.get("detection", "1")

        recommended_action = request.form.get(
            "recommended_action", ""
        ).strip()

        responsibility = request.form.get(
            "responsibility", ""
        ).strip()

        target_date = request.form.get(
            "target_date", ""
        ).strip()

        action_status = request.form.get(
            "action_status", "Open"
        ).strip()

        try:
            s = max(1, min(10, int(severity)))
            o = max(1, min(10, int(occurrence)))
            d = max(1, min(10, int(detection)))

            rpn = s * o * d

        except ValueError:
            s = 1
            o = 1
            d = 1
            rpn = 1

        if project_id and failure_mode:

            conn.execute("""
                INSERT INTO dfmea
                (
                    project_id,
                    component_id,
                    function,
                    failure_mode,
                    failure_effect,
                    severity,
                    cause,
                    occurrence,
                    prevention_control,
                    detection_control,
                    detection,
                    rpn,
                    recommended_action,
                    responsibility,
                    target_date,
                    action_status
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                project_id,
                component_id or None,
                function,
                failure_mode,
                failure_effect,
                s,
                cause,
                o,
                prevention_control,
                detection_control,
                d,
                rpn,
                recommended_action,
                responsibility,
                target_date,
                action_status
            ))

            conn.commit()
            conn.close()

            return redirect(
                url_for(
                    "dfmea",
                    project_id=project_id
                )
            )

    projects = conn.execute("""
        SELECT
            id,
            project_name,
            product_name
        FROM projects
        ORDER BY id DESC
    """).fetchall()

    if selected_project_id:

        components = conn.execute("""
            SELECT
                id,
                component_name,
                component_type,
                part_number,
                level
            FROM product_structure
            WHERE project_id = ?
            ORDER BY level, id
        """, (
            selected_project_id,
        )).fetchall()

    else:
        components = []

    records = conn.execute("""
        SELECT
            d.*,
            p.project_name,
            ps.component_name
        FROM dfmea AS d

        LEFT JOIN projects AS p
            ON d.project_id = p.id

        LEFT JOIN product_structure AS ps
            ON d.component_id = ps.id

        ORDER BY d.id DESC
    """).fetchall()

    conn.close()

    return render_template(
        "dfmea.html",
        projects=projects,
        components=components,
        records=records,
        selected_project_id=selected_project_id
    )


# =========================================================
# PFMEA
# =========================================================

@app.route("/pfmea", methods=["GET", "POST"])
def pfmea():

    conn = get_db()

    selected_project_id = request.args.get(
        "project_id",
        ""
    )

    if request.method == "POST":

        project_id = request.form.get(
            "project_id",
            ""
        )

        component_id = request.form.get(
            "component_id",
            ""
        )

        process_step = request.form.get(
            "process_step",
            ""
        ).strip()

        process_function = request.form.get(
            "process_function",
            ""
        ).strip()

        failure_mode = request.form.get(
            "failure_mode",
            ""
        ).strip()

        failure_effect = request.form.get(
            "failure_effect",
            ""
        ).strip()

        severity = request.form.get(
            "severity",
            "1"
        )

        cause = request.form.get(
            "cause",
            ""
        ).strip()

        occurrence = request.form.get(
            "occurrence",
            "1"
        )

        prevention_control = request.form.get(
            "prevention_control",
            ""
        ).strip()

        detection_control = request.form.get(
            "detection_control",
            ""
        ).strip()

        detection = request.form.get(
            "detection",
            "1"
        )

        recommended_action = request.form.get(
            "recommended_action",
            ""
        ).strip()

        responsibility = request.form.get(
            "responsibility",
            ""
        ).strip()

        target_date = request.form.get(
            "target_date",
            ""
        ).strip()

        action_status = request.form.get(
            "action_status",
            "Open"
        ).strip()

        try:

            s = max(
                1,
                min(
                    10,
                    int(severity)
                )
            )

            o = max(
                1,
                min(
                    10,
                    int(occurrence)
                )
            )

            d = max(
                1,
                min(
                    10,
                    int(detection)
                )
            )

            rpn = s * o * d

        except ValueError:

            s = 1
            o = 1
            d = 1
            rpn = 1

        if project_id and failure_mode:

            conn.execute("""
                INSERT INTO pfmea
                (
                    project_id,
                    component_id,
                    process_step,
                    process_function,
                    failure_mode,
                    failure_effect,
                    severity,
                    cause,
                    occurrence,
                    prevention_control,
                    detection_control,
                    detection,
                    rpn,
                    recommended_action,
                    responsibility,
                    target_date,
                    action_status
                )
                VALUES
                (
                    ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
            """, (
                project_id,
                component_id or None,
                process_step,
                process_function,
                failure_mode,
                failure_effect,
                s,
                cause,
                o,
                prevention_control,
                detection_control,
                d,
                rpn,
                recommended_action,
                responsibility,
                target_date,
                action_status
            ))

            conn.commit()
            conn.close()

            return redirect(
                url_for(
                    "pfmea",
                    project_id=project_id
                )
            )

    projects = conn.execute("""
        SELECT
            id,
            project_name,
            product_name
        FROM projects
        ORDER BY id DESC
    """).fetchall()

    if selected_project_id:

        components = conn.execute("""
            SELECT
                id,
                component_name,
                component_type,
                part_number,
                level
            FROM product_structure
            WHERE project_id = ?
            ORDER BY level, id
        """, (
            selected_project_id,
        )).fetchall()

    else:

        components = []

    records = conn.execute("""
        SELECT
            p.*,
            pr.project_name,
            ps.component_name

        FROM pfmea AS p

        LEFT JOIN projects AS pr
            ON p.project_id = pr.id

        LEFT JOIN product_structure AS ps
            ON p.component_id = ps.id

        ORDER BY p.id DESC
    """).fetchall()

    conn.close()

    return render_template(
        "pfmea.html",
        projects=projects,
        components=components,
        records=records,
        selected_project_id=selected_project_id
    )


# =========================================================
# CONTROL PLAN
# =========================================================

@app.route("/control-plan", methods=["GET", "POST"])
def control_plan():

    conn = get_db()

    selected_project_id = request.args.get(
        "project_id",
        ""
    )

    if request.method == "POST":

        project_id = request.form.get(
            "project_id",
            ""
        )

        component_id = request.form.get(
            "component_id",
            ""
        )

        process_step = request.form.get(
            "process_step",
            ""
        ).strip()

        characteristic = request.form.get(
            "characteristic",
            ""
        ).strip()

        specification = request.form.get(
            "specification",
            ""
        ).strip()

        control_method = request.form.get(
            "control_method",
            ""
        ).strip()

        measurement_method = request.form.get(
            "measurement_method",
            ""
        ).strip()

        sample_size = request.form.get(
            "sample_size",
            ""
        ).strip()

        frequency = request.form.get(
            "frequency",
            ""
        ).strip()

        responsibility = request.form.get(
            "responsibility",
            ""
        ).strip()

        reaction_plan = request.form.get(
            "reaction_plan",
            ""
        ).strip()

        if project_id and characteristic:

            conn.execute("""
                INSERT INTO control_plan
                (
                    project_id,
                    component_id,
                    process_step,
                    characteristic,
                    specification,
                    control_method,
                    measurement_method,
                    sample_size,
                    frequency,
                    responsibility,
                    reaction_plan
                )
                VALUES
                (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                project_id,
                component_id or None,
                process_step,
                characteristic,
                specification,
                control_method,
                measurement_method,
                sample_size,
                frequency,
                responsibility,
                reaction_plan
            ))

            conn.commit()
            conn.close()

            return redirect(
                url_for(
                    "control_plan",
                    project_id=project_id
                )
            )

    projects = conn.execute("""
        SELECT
            id,
            project_name,
            product_name
        FROM projects
        ORDER BY id DESC
    """).fetchall()

    if selected_project_id:

        components = conn.execute("""
            SELECT
                id,
                component_name,
                component_type,
                part_number,
                level
            FROM product_structure
            WHERE project_id = ?
            ORDER BY level, id
        """, (
            selected_project_id,
        )).fetchall()

    else:

        components = []

    records = conn.execute("""
        SELECT
            cp.*,
            p.project_name,
            ps.component_name

        FROM control_plan AS cp

        LEFT JOIN projects AS p
            ON cp.project_id = p.id

        LEFT JOIN product_structure AS ps
            ON cp.component_id = ps.id

        ORDER BY cp.id DESC
    """).fetchall()

    conn.close()

    return render_template(
        "control_plan.html",
        projects=projects,
        components=components,
        records=records,
        selected_project_id=selected_project_id
    )


# =========================================================
# REPORTS
# =========================================================

@app.route("/reports", methods=["GET"])
def reports():

    conn = get_db()

    selected_project_id = request.args.get(
        "project_id",
        ""
    )

    projects = conn.execute("""
        SELECT *
        FROM projects
        ORDER BY id DESC
    """).fetchall()

    project_info = None

    summary = {
        "functional_analysis": 0,
        "boundary_diagram": 0,
        "product_structure": 0,
        "key_characteristics": 0,
        "functional_links": 0,
        "dfmea": 0,
        "pfmea": 0,
        "control_plan": 0
    }

    components = []
    dfmea_records = []
    pfmea_records = []

    if selected_project_id:

        project_info = conn.execute("""
            SELECT *
            FROM projects
            WHERE id = ?
        """, (
            selected_project_id,
        )).fetchone()

        tables = [
            "functional_analysis",
            "boundary_diagram",
            "product_structure",
            "key_characteristics",
            "functional_links",
            "dfmea",
            "pfmea",
            "control_plan"
        ]

        for table in tables:

            summary[table] = conn.execute(
                f"""
                SELECT COUNT(*)
                FROM {table}
                WHERE project_id = ?
                """,
                (
                    selected_project_id,
                )
            ).fetchone()[0]

        components = conn.execute("""
            SELECT
                id,
                component_name,
                component_type,
                label,
                part_number,
                level
            FROM product_structure
            WHERE project_id = ?
            ORDER BY level, id
        """, (
            selected_project_id,
        )).fetchall()

        dfmea_records = conn.execute("""
            SELECT
                d.*,
                ps.component_name

            FROM dfmea AS d

            LEFT JOIN product_structure AS ps
                ON d.component_id = ps.id

            WHERE d.project_id = ?

            ORDER BY d.id DESC
        """, (
            selected_project_id,
        )).fetchall()

        pfmea_records = conn.execute("""
            SELECT
                p.*,
                ps.component_name

            FROM pfmea AS p

            LEFT JOIN product_structure AS ps
                ON p.component_id = ps.id

            WHERE p.project_id = ?

            ORDER BY p.id DESC
        """, (
            selected_project_id,
        )).fetchall()

    conn.close()

    return render_template(
        "reports.html",
        projects=projects,
        project_info=project_info,
        summary=summary,
        components=components,
        dfmea_records=dfmea_records,
        pfmea_records=pfmea_records,
        selected_project_id=selected_project_id
    )


# =========================================================
# PROFESSIONAL EXCEL HELPERS
# =========================================================

def apply_sheet_format(
    sheet,
    title,
    subtitle,
    landscape=True,
    tab_color="2F75B5"
):

    dark_blue = "17365D"
    medium_blue = "2F75B5"
    border_color = "B7C9D6"

    thin = Side(
        style="thin",
        color=border_color
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    sheet.sheet_view.showGridLines = False

    sheet.sheet_properties.tabColor = tab_color

    max_col = max(
        sheet.max_column,
        1
    )

    # TITLE
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max_col
    )

    title_cell = sheet.cell(
        1,
        1
    )

    title_cell.value = title

    title_cell.font = Font(
        name="Aptos",
        size=18,
        bold=True,
        color="FFFFFF"
    )

    title_cell.fill = PatternFill(
        "solid",
        fgColor=dark_blue
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    sheet.row_dimensions[1].height = 32

    # SUBTITLE
    sheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=max_col
    )

    subtitle_cell = sheet.cell(
        2,
        1
    )

    subtitle_cell.value = subtitle

    subtitle_cell.font = Font(
        name="Aptos",
        size=10,
        italic=True
    )

    subtitle_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    sheet.row_dimensions[2].height = 22

    sheet.row_dimensions[3].height = 8

    # HEADER
    for cell in sheet[4]:

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=medium_blue
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = border

    sheet.row_dimensions[4].height = 34

    # DATA
    for row_number in range(
        5,
        sheet.max_row + 1
    ):

        sheet.row_dimensions[
            row_number
        ].height = 40

        for cell in sheet[row_number]:

            cell.font = Font(
                name="Aptos",
                size=10
            )

            cell.border = border

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            if row_number % 2 == 1:

                cell.fill = PatternFill(
                    "solid",
                    fgColor="F4F8FB"
                )

    # FILTER
    if sheet.max_row >= 4:

        sheet.auto_filter.ref = (
            f"A4:"
            f"{get_column_letter(sheet.max_column)}"
            f"{sheet.max_row}"
        )

    # FREEZE
    sheet.freeze_panes = "A5"

    # WIDTH
    for column in range(
        1,
        sheet.max_column + 1
    ):

        letter = get_column_letter(
            column
        )

        maximum = 0

        for row in range(
            1,
            sheet.max_row + 1
        ):

            value = sheet.cell(
                row,
                column
            ).value

            if value is not None:

                maximum = max(
                    maximum,
                    len(str(value))
                )

        sheet.column_dimensions[
            letter
        ].width = min(
            max(
                maximum + 3,
                12
            ),
            35
        )

    # PRINT SETTINGS
    sheet.page_setup.orientation = (
        "landscape"
        if landscape
        else "portrait"
    )

    sheet.page_setup.paperSize = (
        sheet.PAPERSIZE_A4
    )

    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    sheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2
    )

    sheet.print_title_rows = "1:4"

    sheet.oddFooter.center.text = (
        "Automotive FMEA Management System"
    )

    sheet.oddFooter.right.text = (
        "Page &P of &N"
    )


def add_rpn_rules(
    sheet,
    column_number
):

    if sheet.max_row < 5:
        return

    letter = get_column_letter(
        column_number
    )

    cell_range = (
        f"{letter}5:"
        f"{letter}{sheet.max_row}"
    )

    # HIGH RPN
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["200"],
            fill=PatternFill(
                "solid",
                fgColor="F4CCCC"
            )
        )
    )

    # MEDIUM RPN
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="between",
            formula=[
                "100",
                "199"
            ],
            fill=PatternFill(
                "solid",
                fgColor="FFF2CC"
            )
        )
    )

    # LOW RPN
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThan",
            formula=["100"],
            fill=PatternFill(
                "solid",
                fgColor="D9EAD3"
            )
        )
    )


# =========================================================
# EXCEL EXPORT
# =========================================================

@app.route("/export-excel")
def export_excel():

    project_id = request.args.get(
        "project_id"
    )

    if not project_id:
        return "Please select a project first."

    conn = get_db()

    project = conn.execute("""
        SELECT *
        FROM projects
        WHERE id = ?
    """, (
        project_id,
    )).fetchone()

    if not project:

        conn.close()

        return "Project not found."

    workbook = Workbook()

    # =====================================================
    # PROJECT SHEET
    # =====================================================

    sheet = workbook.active

    sheet.title = "Project"

    sheet.cell(
        4,
        1,
        "Field"
    )

    sheet.cell(
        4,
        2,
        "Project Details"
    )

    project_rows = [
        (
            "Project Name",
            project["project_name"]
        ),
        (
            "Product Name",
            project["product_name"]
        ),
        (
            "Customer",
            project["customer"]
        ),
        (
            "OEM / Customer Standard",
            project["oem_name"]
        ),
        (
            "Compliance Mode",
            project["compliance_mode"]
        ),
        (
            "Project Number",
            project["project_number"]
        ),
        (
            "Created Date",
            project["created_date"]
        )
    ]

    for row_number, (
        label,
        value
    ) in enumerate(
        project_rows,
        start=5
    ):

        sheet.cell(
            row_number,
            1,
            label
        )

        sheet.cell(
            row_number,
            2,
            value
        )

    apply_sheet_format(
        sheet,
        "AUTOMOTIVE FMEA MANAGEMENT SYSTEM",
        "Project Identification & Configuration",
        landscape=False,
        tab_color="17365D"
    )

    sheet.column_dimensions[
        "A"
    ].width = 30

    sheet.column_dimensions[
        "B"
    ].width = 50

    for row_number in range(
        5,
        12
    ):

        sheet.cell(
            row_number,
            1
        ).font = Font(
            name="Aptos",
            size=10,
            bold=True
        )

        sheet.cell(
            row_number,
            1
        ).fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )


    # =====================================================
    # PRODUCT STRUCTURE
    # =====================================================

    sheet = workbook.create_sheet(
        "Product Structure"
    )

    sheet.append([
        "Level",
        "Component",
        "Type",
        "Label",
        "Part Number",
        "Description"
    ])

    rows = conn.execute("""
        SELECT
            level,
            component_name,
            component_type,
            label,
            part_number,
            description
        FROM product_structure
        WHERE project_id = ?
        ORDER BY level, id
    """, (
        project_id,
    )).fetchall()

    for row in rows:

        sheet.append([
            row["level"],
            row["component_name"],
            row["component_type"],
            row["label"],
            row["part_number"],
            row["description"]
        ])

    apply_sheet_format(
        sheet,
        "PRODUCT STRUCTURE",
        "Product hierarchy and component definition",
        True,
        "5B9BD5"
    )


    # =====================================================
    # FUNCTIONAL ANALYSIS
    # =====================================================

    sheet = workbook.create_sheet(
        "Functional Analysis"
    )

    sheet.append([
        "Level",
        "Function",
        "Requirement"
    ])

    rows = conn.execute("""
        SELECT
            function,
            requirement
        FROM functional_analysis
        WHERE project_id = ?
        ORDER BY id
    """, (
        project_id,
    )).fetchall()

    for row in rows:

        sheet.append([
            "",
            row["function"],
            row["requirement"]
        ])

    apply_sheet_format(
        sheet,
        "FUNCTIONAL ANALYSIS",
        "Product functions, levels and associated requirements",
        True,
        "70AD47"
    )


    # =====================================================
    # BOUNDARY DIAGRAM
    # =====================================================

    sheet = workbook.create_sheet(
        "Boundary Diagram"
    )

    sheet.append([
        "External Element",
        "Interaction",
        "Direction",
        "Description"
    ])

    rows = conn.execute("""
        SELECT
            external_element,
            interaction,
            direction,
            description
        FROM boundary_diagram
        WHERE project_id = ?
        ORDER BY id
    """, (
        project_id,
    )).fetchall()

    for row in rows:

        sheet.append([
            row["external_element"],
            row["interaction"],
            row["direction"],
            row["description"]
        ])

    apply_sheet_format(
        sheet,
        "BOUNDARY DIAGRAM",
        "System boundaries and external interactions",
        True,
        "ED7D31"
    )


    # =====================================================
    # KEY CHARACTERISTICS
    # =====================================================

    sheet = workbook.create_sheet(
        "Key Characteristics"
    )

    sheet.append([
        "Component",
        "Characteristic",
        "Specification",
        "Tolerance",
        "Severity",
        "Responsibility"
    ])

    rows = conn.execute("""
        SELECT
            ps.component_name,
            kc.characteristic,
            kc.specification,
            kc.tolerance,
            kc.severity,
            kc.responsibility

        FROM key_characteristics AS kc

        LEFT JOIN product_structure AS ps
            ON kc.component_id = ps.id

        WHERE kc.project_id = ?

        ORDER BY kc.id
    """, (
        project_id,
    )).fetchall()

    for row in rows:

        sheet.append([
            row["component_name"],
            row["characteristic"],
            row["specification"],
            row["tolerance"],
            row["severity"],
            row["responsibility"]
        ])

    apply_sheet_format(
        sheet,
        "KEY CHARACTERISTICS",
        "Special characteristics and specifications",
        True,
        "FFC000"
    )


    # =====================================================
    # FUNCTIONAL LINKS
    # =====================================================

    sheet = workbook.create_sheet(
        "Functional Links"
    )

    sheet.append([
        "Function",
        "Function Requirement",
        "Component",
        "Linked Requirement"
    ])

    rows = conn.execute("""
        SELECT
            fa.function AS function_name,
            fa.requirement AS function_requirement,
            ps.component_name,
            fl.requirement AS linked_requirement

        FROM functional_links AS fl

        LEFT JOIN functional_analysis AS fa
            ON fl.function_id = fa.id

        LEFT JOIN product_structure AS ps
            ON fl.component_id = ps.id

        WHERE fl.project_id = ?

        ORDER BY fl.id
    """, (
        project_id,
    )).fetchall()

    for row in rows:

        sheet.append([
            row["function_name"],
            row["function_requirement"],
            row["component_name"],
            row["linked_requirement"]
        ])

    apply_sheet_format(
        sheet,
        "FUNCTIONAL LINKING",
        "Traceability between functions, requirements and components",
        True,
        "A5A5A5"
    )


    # =====================================================
    # DFMEA
    # =====================================================

    sheet = workbook.create_sheet(
        "DFMEA"
    )

    sheet.append([
        "Component",
        "Function",
        "Failure Mode",
        "Failure Effect",
        "Severity",
        "Potential Cause",
        "Occurrence",
        "Prevention Control",
        "Detection Control",
        "Detection",
        "RPN",
        "Recommended Action",
        "Responsibility",
        "Target Date",
        "Action Status"
    ])

    rows = conn.execute("""
        SELECT
            ps.component_name,
            d.function,
            d.failure_mode,
            d.failure_effect,
            d.severity,
            d.cause,
            d.occurrence,
            d.prevention_control,
            d.detection_control,
            d.detection,
            d.rpn,
            d.recommended_action,
            d.responsibility,
            d.target_date,
            d.action_status

        FROM dfmea AS d

        LEFT JOIN product_structure AS ps
            ON d.component_id = ps.id

        WHERE d.project_id = ?

        ORDER BY d.id
    """, (
        project_id,
    )).fetchall()

    for row in rows:

        sheet.append([
            row["component_name"],
            row["function"],
            row["failure_mode"],
            row["failure_effect"],
            row["severity"],
            row["cause"],
            row["occurrence"],
            row["prevention_control"],
            row["detection_control"],
            row["detection"],
            row["rpn"],
            row["recommended_action"],
            row["responsibility"],
            row["target_date"],
            row["action_status"]
        ])

    apply_sheet_format(
        sheet,
        "DESIGN FMEA (DFMEA)",
        "Design failure mode and effects analysis",
        True,
        "4472C4"
    )

    add_rpn_rules(
        sheet,
        11
    )


    # =====================================================
    # PFMEA
    # =====================================================

    sheet = workbook.create_sheet(
        "PFMEA"
    )

    sheet.append([
        "Component",
        "Process Step",
        "Process Function",
        "Failure Mode",
        "Failure Effect",
        "Severity",
        "Potential Cause",
        "Occurrence",
        "Prevention Control",
        "Detection Control",
        "Detection",
        "RPN",
        "Recommended Action",
        "Responsibility",
        "Target Date",
        "Action Status"
    ])

    rows = conn.execute("""
        SELECT
            ps.component_name,
            p.process_step,
            p.process_function,
            p.failure_mode,
            p.failure_effect,
            p.severity,
            p.cause,
            p.occurrence,
            p.prevention_control,
            p.detection_control,
            p.detection,
            p.rpn,
            p.recommended_action,
            p.responsibility,
            p.target_date,
            p.action_status

        FROM pfmea AS p

        LEFT JOIN product_structure AS ps
            ON p.component_id = ps.id

        WHERE p.project_id = ?

        ORDER BY p.id
    """, (
        project_id,
    )).fetchall()

    for row in rows:

        sheet.append([
            row["component_name"],
            row["process_step"],
            row["process_function"],
            row["failure_mode"],
            row["failure_effect"],
            row["severity"],
            row["cause"],
            row["occurrence"],
            row["prevention_control"],
            row["detection_control"],
            row["detection"],
            row["rpn"],
            row["recommended_action"],
            row["responsibility"],
            row["target_date"],
            row["action_status"]
        ])

    apply_sheet_format(
        sheet,
        "PROCESS FMEA (PFMEA)",
        "Process failure mode and effects analysis",
        True,
        "70AD47"
    )

    add_rpn_rules(
        sheet,
        12
    )


    # =====================================================
    # CONTROL PLAN
    # =====================================================

    sheet = workbook.create_sheet(
        "Control Plan"
    )

    sheet.append([
        "Component",
        "Process Step",
        "Characteristic",
        "Specification",
        "Control Method",
        "Measurement Method",
        "Sample Size",
        "Frequency",
        "Responsibility",
        "Reaction Plan"
    ])

    rows = conn.execute("""
        SELECT
            ps.component_name,
            cp.process_step,
            cp.characteristic,
            cp.specification,
            cp.control_method,
            cp.measurement_method,
            cp.sample_size,
            cp.frequency,
            cp.responsibility,
            cp.reaction_plan

        FROM control_plan AS cp

        LEFT JOIN product_structure AS ps
            ON cp.component_id = ps.id

        WHERE cp.project_id = ?

        ORDER BY cp.id
    """, (
        project_id,
    )).fetchall()

    for row in rows:

        sheet.append([
            row["component_name"],
            row["process_step"],
            row["characteristic"],
            row["specification"],
            row["control_method"],
            row["measurement_method"],
            row["sample_size"],
            row["frequency"],
            row["responsibility"],
            row["reaction_plan"]
        ])

    apply_sheet_format(
        sheet,
        "CONTROL PLAN",
        "Process controls, measurements and reaction plans",
        True,
        "5B9BD5"
    )

    conn.close()


    # =====================================================
    # SAVE EXCEL
    # =====================================================

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Automotive_FMEA_Report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# =========================================================
# INITIALIZE DATABASE
# =========================================================

setup_database()


# =========================================================
# START APPLICATION
# =========================================================

if __name__ == "__main__":

    print("")
    print("==============================================")
    print("      AUTOMOTIVE FMEA MANAGEMENT SYSTEM")
    print("==============================================")

    app.run(
        debug=False,
        host="0.0.0.0",
        port=int(
            os.environ.get(
                "PORT",
                5000
            )
        )
    )
